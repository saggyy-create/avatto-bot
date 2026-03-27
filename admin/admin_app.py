import os
import io
import json
import shutil
import logging
from pathlib import Path
from typing import Optional

import openpyxl
import aiosqlite
from fastapi import FastAPI, Request, Form, File, UploadFile, Depends, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

import sys
sys.path.insert(0, str(Path(__file__).parent.parent / "bot"))

from config import DATABASE_PATH, UPLOAD_DIR, ADMIN_IDS
from database.db import (
    get_all_products, get_all_users, get_all_orders,
    create_product, update_product, delete_product, get_product,
    bulk_create_products
)

logging.basicConfig(level=logging.INFO)
app = FastAPI(title="Catalog Bot Admin")

UPLOAD_PATH = Path(UPLOAD_DIR)
UPLOAD_PATH.mkdir(exist_ok=True)

BASE_DIR = Path(__file__).parent
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")
app.mount("/uploads", StaticFiles(directory=str(UPLOAD_PATH)), name="uploads")


# ─── Products API (used by Mini App) ──────────────────────────────────────────

@app.get("/api/products")
async def api_products():
    products = await get_all_products()
    result = []
    for p in products:
        result.append({
            "id": p["id"],
            "name": p["name"],
            "description": p["description"],
            "price": p["price"],
            "article": p["article"],
            "photo_url": f"/uploads/{p['photo_path']}" if p["photo_path"] else None,
        })
    return JSONResponse(result)


# ─── Admin Pages ───────────────────────────────────────────────────────────────

@app.get("/admin", response_class=HTMLResponse)
async def admin_index(request: Request):
    products = await get_all_products()
    users = await get_all_users()
    orders = await get_all_orders()
    return templates.TemplateResponse("index.html", {
        "request": request,
        "products_count": len(products),
        "users_count": len(users),
        "orders_count": len(orders),
    })


@app.get("/admin/products", response_class=HTMLResponse)
async def admin_products(request: Request):
    products = await get_all_products()
    return templates.TemplateResponse("products.html", {
        "request": request, "products": products
    })


@app.post("/admin/products/add")
async def admin_add_product(
    request: Request,
    name: str = Form(...),
    description: str = Form(""),
    price: str = Form(""),
    article: str = Form(""),
    photo: UploadFile = File(None)
):
    photo_path = None
    if photo and photo.filename:
        suffix = Path(photo.filename).suffix
        filename = f"product_{name.replace(' ', '_')[:20]}{suffix}"
        filepath = UPLOAD_PATH / filename
        with open(filepath, "wb") as f:
            shutil.copyfileobj(photo.file, f)
        photo_path = filename

    price_val = float(price.strip().replace('$','').replace(',','.')) if price.strip() else None
```
    await create_product(name, description or None, price_val, article or None, photo_path)
    return RedirectResponse("/admin/products", status_code=303)


@app.get("/admin/products/{product_id}/edit", response_class=HTMLResponse)
async def admin_edit_product_form(request: Request, product_id: int):
    product = await get_product(product_id)
    if not product:
        raise HTTPException(404, "Product not found")
    return templates.TemplateResponse("edit_product.html", {
        "request": request, "product": product
    })


@app.post("/admin/products/{product_id}/edit")
async def admin_edit_product(
    product_id: int,
    name: str = Form(...),
    description: str = Form(""),
    price: str = Form(""),
    article: str = Form(""),
    photo: UploadFile = File(None)
):
    existing = await get_product(product_id)
    photo_path = existing["photo_path"] if existing else None

    if photo and photo.filename:
        suffix = Path(photo.filename).suffix
        filename = f"product_{product_id}{suffix}"
        filepath = UPLOAD_PATH / filename
        with open(filepath, "wb") as f:
            shutil.copyfileobj(photo.file, f)
        photo_path = filename

    price_val = float(price) if price.strip() else None
    await update_product(product_id, name, description or None, price_val, article or None, photo_path)
    return RedirectResponse("/admin/products", status_code=303)


@app.post("/admin/products/{product_id}/delete")
async def admin_delete_product(product_id: int):
    await delete_product(product_id)
    return RedirectResponse("/admin/products", status_code=303)


@app.post("/admin/products/import")
async def admin_import_products(file: UploadFile = File(...)):
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    products = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row[0]:
            continue
        products.append({
            "name": str(row[0]),
            "description": str(row[1]) if row[1] else None,
            "price": float(row[2]) if row[2] else None,
            "article": str(row[3]) if row[3] else None,
        })

    await bulk_create_products(products)
    return RedirectResponse(f"/admin/products?imported={len(products)}", status_code=303)


@app.get("/admin/excel-template")
async def download_excel_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Товары"
    ws.append(["Название", "Описание", "Цена", "Артикул"])
    ws.append(["Пример товара 1", "Описание товара", "1500", "ART-001"])
    ws.append(["Пример товара 2", "Ещё один товар", "2300", "ART-002"])

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15

    path = "/tmp/import_template.xlsx"
    wb.save(path)
    return FileResponse(path, filename="import_template.xlsx",
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/admin/users", response_class=HTMLResponse)
async def admin_users(request: Request):
    users = await get_all_users()
    return templates.TemplateResponse("users.html", {
        "request": request, "users": users
    })


@app.get("/admin/orders", response_class=HTMLResponse)
async def admin_orders(request: Request, search: str = ""):
    orders = await get_all_orders()
    
    # Parse product ids per order and attach names
    all_products = {p["id"]: p for p in await get_all_products()}
    
    enriched = []
    for o in orders:
        product_ids = json.loads(o["products"])
        product_names = [all_products[pid]["name"] for pid in product_ids if pid in all_products]
        enriched.append({
            "id": o["id"],
            "user_name": o["user_name"],
            "phone": o["phone"],
            "telegram_id": o["telegram_id"],
            "products": product_names,
            "created_at": o["created_at"],
        })

    if search:
        s = search.lower()
        enriched = [o for o in enriched if s in o["user_name"].lower() or s in o["phone"]]

    return templates.TemplateResponse("orders.html", {
        "request": request, "orders": enriched, "search": search
    })


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
